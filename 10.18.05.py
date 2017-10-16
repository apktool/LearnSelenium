from openpyxl import Workbook
import mysql.connector
import time

wb = Workbook()
ws = wb.active


def execute_sql(sql):
    try:
        cnx = mysql.connector.connect(user='python', password='python', database='testDB')
    except mysql.connector.Error as err:
        print("Connetion errors have done {}".format(err))
    finally:
        cursor = cnx.cursor()

    cursor.execute(sql)
    # return cursor
    # cnx.commit()


    for account, top_title, all_names, all_read_counts, all_links_counts, all_links, all_times in cursor:
        item = []
        item.append(account)
        item.append(top_title)
        item.append(all_names)
        item.append(all_read_counts)
        item.append(all_links_counts)
        item.append(all_links)
        item.append(str(all_times))
        write_to_xlsx(item)


def write_to_xlsx(item):
    ws['A1'] = 'Account'
    ws['B1'] = 'Popularity/Yesterday'
    ws['C1'] = 'Title'
    ws['D1'] = 'Reading'
    ws['E1'] = 'Praise'
    ws['F1'] = 'Ticket'
    ws['G1'] = 'DateTime'

    ws.append(item)
    wb.save("sample.xlsx")


if __name__ == '__main__':
    sql = "SELECT account, top_title, all_names, all_read_counts, all_links_counts, all_links, all_times FROM newrank_account_info;"
    cursor = execute_sql(sql)
