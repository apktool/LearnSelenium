from openpyxl import Workbook
import mysql.connector
import time

wb = Workbook()
ws = wb.active


def execute_sql(sql):
    try:
        cnx = mysql.connector.connect(user='python', password='python', database='testDB')
    except mysql.connector.Error as err:
        print("Connetion errors have done {}".format(err))
    finally:
        cursor = cnx.cursor()

    cursor.execute(sql)
    # return cursor
    # cnx.commit()


    from account import accounts
    accounts = {v: k for k, v in accounts.items()}
    for account, top_title, title, summary, all_read_counts, all_links_counts, link, publish_time in cursor:
        item = []
        item.append(accounts[account])
        item.append(top_title)
        item.append(title)
        item.append(summary)
        item.append(all_read_counts)
        item.append(all_links_counts)
        item.append(link)
        item.append(str(publish_time))
        write_to_xlsx(item)


def write_to_xlsx(item):
    ws['A1'] = 'Account'
    ws['B1'] = 'Popularity/Yesterday'
    ws['C1'] = 'Title'
    ws['D1'] = 'Summary'
    ws['E1'] = 'Reading'
    ws['F1'] = 'Praise'
    ws['G1'] = 'Ticket'
    ws['H1'] = 'publish_time'

    ws.append(item)
    wb.save("sample.xlsx")


if __name__ == '__main__':
    sql = "SELECT account, top_title, title, summary, all_read_counts, all_links_counts, link, publish_time FROM newrank_account_info;"
    cursor = execute_sql(sql)
